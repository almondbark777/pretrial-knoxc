"""
build_db.py
-----------
Builds the KH222 Knox-County Pre-Trial SQLite database from the five raw
SharePoint/JIMS source files:

    KH222 Master List.xlsx      (ALL sheet, 8 cols)
    New Blue Book.csv           (40 cols)
    Payments (2).csv            (8 cols)
    Check Ins (1).csv           (9 cols)
    GPS 48 Hours (1).csv        (19 cols)

Outputs:
    kh222.db    - SQLite with raw_* tables + normalized tables + view

Design notes:
  * CSVs from SharePoint's "Export to CSV" have a giant ListSchema JSON
    blob as row 0.  We strip row 0 before parsing.  Encoding is utf-8-sig.
  * Column names snake_case'd, %23 -> 'num', collisions suffixed _2,_3,...
  * Dates try ISO first, then %m/%d/%Y %H:%M, %m/%d/%Y, %Y-%m-%d %H:%M:%S,
    %Y-%m-%d; unparseable strings kept raw.
  * Currency "$20.00" -> 20.0 float.
  * Booleans True/False/Yes/No -> 0/1 ints.
  * Case numbers "@1632112" or "@1606962, @1641152": raw_* tables keep the
    string; normalized `cases` table splits and strips the '@'.

Environment gotcha: the OneDrive-mounted outputs folder doesn't support
SQLite file locking.  Build in /tmp, then byte-copy the finished .db to
the final destination.

Usage:
    python build_db.py --uploads <folder> --out <folder>
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl required: pip install openpyxl")


# ----------------------------------------------------------------
# Column / value normalization helpers
# ----------------------------------------------------------------

def snake(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().replace("%23", "num")
    s = re.sub(r"[^\w]+", "_", s).strip("_")
    s = re.sub(r"_+", "_", s)
    return s.lower()


def dedupe_columns(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 1
            out.append(c)
    return out


_DATE_FORMATS = (
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
)


def parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%dT%H:%M:%S")
    s = str(v).strip()
    if not s:
        return None
    # ISO w/ Z
    try:
        if s.endswith("Z"):
            dt = datetime.strptime(s[:-1], "%Y-%m-%dT%H:%M:%S")
            return dt.strftime("%Y-%m-%dT%H:%M:%S")
        if "T" in s:
            return datetime.fromisoformat(s).strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%S")
        except Exception:
            continue
    return s  # leave as-is


_MONEY_RE = re.compile(r"[^\d\.\-]")


def parse_money(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _MONEY_RE.sub("", str(v))
    if not s or s in {".", "-"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_bool(v):
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return int(v)
    s = str(v).strip().lower()
    if s in {"true", "yes", "1", "y", "t"}:
        return 1
    if s in {"false", "no", "0", "n", "f"}:
        return 0
    return None


def parse_int(v):
    if v is None or v == "":
        return None
    if isinstance(v, int):
        return v
    try:
        return int(float(str(v).replace(",", "")))
    except ValueError:
        return None


def split_cases(raw) -> list[str]:
    """@1606962, @1641152  ->  ['1606962', '1641152']"""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    out = []
    for piece in re.split(r"[,;\s]+", s):
        p = piece.strip().lstrip("@").strip()
        if p and p not in out:
            out.append(p)
    return out


# ----------------------------------------------------------------
# Source readers
# ----------------------------------------------------------------

def read_sharepoint_csv(path: Path) -> tuple[list[str], list[list]]:
    """Strip the ListSchema JSON row-0 blob and read rows."""
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        rdr = csv.reader(fh)
        rows = list(rdr)
    if not rows:
        return [], []
    # Row 0 is the ListSchema blob; Row 1 is the real header.
    header = rows[1] if len(rows) > 1 else []
    data = rows[2:] if len(rows) > 2 else []
    header = dedupe_columns([snake(h) for h in header])
    return header, data


def read_master_list(path: Path) -> tuple[list[str], list[list]]:
    """Read only the ALL sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ALL" not in wb.sheetnames:
        raise RuntimeError(f"ALL sheet not found in {path.name}")
    ws = wb["ALL"]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    header = dedupe_columns([snake(h) for h in header])
    data = [list(r) for r in it]
    return header, data


# ----------------------------------------------------------------
# Schema
# ----------------------------------------------------------------

DDL = [
    """CREATE TABLE defendants (
        idn INTEGER PRIMARY KEY,
        defendant_name TEXT,
        defendant_last_name TEXT,
        birthdate TEXT,
        pretrial_level INTEGER,
        charge_type TEXT,
        supervision_type TEXT,
        order_from TEXT,
        dma INTEGER,
        gps INTEGER,
        referral_date TEXT,
        closed_date TEXT,
        case_status TEXT,
        supervising_officer TEXT,
        ptr_successfully_completed INTEGER,
        bond_amount REAL,
        total_paid REAL,
        source TEXT
    )""",
    """CREATE TABLE cases (
        case_id INTEGER PRIMARY KEY AUTOINCREMENT,
        idn INTEGER,
        case_number TEXT NOT NULL,
        source TEXT,
        FOREIGN KEY (idn) REFERENCES defendants(idn)
    )""",
    """CREATE TABLE payments (
        payment_id INTEGER PRIMARY KEY AUTOINCREMENT,
        idn INTEGER,
        case_number TEXT,
        defendant TEXT,
        payment_date TEXT,
        payment_amount REAL,
        officer TEXT,
        payment_type TEXT,
        case_status TEXT,
        FOREIGN KEY (idn) REFERENCES defendants(idn)
    )""",
    """CREATE TABLE check_ins (
        check_in_id INTEGER PRIMARY KEY AUTOINCREMENT,
        idn INTEGER,
        case_number TEXT,
        defendant TEXT,
        check_in_date TEXT,
        type_of_check_in TEXT,
        supervising_officer TEXT,
        case_status TEXT,
        referral_date TEXT,
        pretrial_level INTEGER,
        FOREIGN KEY (idn) REFERENCES defendants(idn)
    )""",
    """CREATE TABLE gps_events (
        gps_id INTEGER PRIMARY KEY AUTOINCREMENT,
        idn INTEGER,
        case_number TEXT,
        defendant TEXT,
        referral_date TEXT,
        gps_type TEXT,
        case_status TEXT,
        paid INTEGER,
        victim TEXT,
        victim_idn INTEGER,
        victim_time_48 TEXT,
        victim_accept_deny_gps INTEGER,
        gps_install_date TEXT,
        court_order TEXT,
        da_emailed INTEGER,
        closed_date TEXT,
        FOREIGN KEY (idn) REFERENCES defendants(idn)
    )""",
    "CREATE INDEX idx_cases_idn       ON cases(idn)",
    "CREATE INDEX idx_cases_case      ON cases(case_number)",
    "CREATE INDEX idx_payments_idn    ON payments(idn)",
    "CREATE INDEX idx_payments_date   ON payments(payment_date)",
    "CREATE INDEX idx_checkins_idn    ON check_ins(idn)",
    "CREATE INDEX idx_checkins_date   ON check_ins(check_in_date)",
    "CREATE INDEX idx_gps_idn         ON gps_events(idn)",
    """CREATE VIEW v_defendant_summary AS
       SELECT d.idn, d.defendant_name, d.defendant_last_name, d.pretrial_level,
              d.charge_type, d.case_status, d.supervising_officer,
              d.referral_date, d.closed_date,
              (SELECT COUNT(*) FROM cases c      WHERE c.idn=d.idn) AS case_count,
              (SELECT COUNT(*) FROM check_ins ci WHERE ci.idn=d.idn) AS check_in_count,
              (SELECT COUNT(*) FROM payments p   WHERE p.idn=d.idn) AS payment_count,
              (SELECT COALESCE(SUM(p.payment_amount),0) FROM payments p WHERE p.idn=d.idn) AS total_paid_calc,
              (SELECT COUNT(*) FROM gps_events g WHERE g.idn=d.idn) AS gps_events_count
       FROM defendants d""",
]


def build_raw_table(con, name, header, rows):
    """Create a raw_<name> table column-for-column and bulk insert."""
    cols_sql = ",\n    ".join(f'"{h}" TEXT' for h in header)
    con.execute(f'CREATE TABLE "{name}" (\n    {cols_sql}\n)')
    if not rows:
        return
    placeholders = ",".join("?" * len(header))
    qcols = ",".join(f'"{h}"' for h in header)
    con.executemany(
        f'INSERT INTO "{name}" ({qcols}) VALUES ({placeholders})',
        rows,
    )


# ----------------------------------------------------------------
# Main ETL
# ----------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--uploads", required=True, help="Folder with source files")
    ap.add_argument("--out",     required=True, help="Folder to drop kh222.db in")
    args = ap.parse_args()

    up = Path(args.uploads).resolve()
    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # Stage the .db in /tmp - OneDrive can't lock SQLite.
    stage = Path("/tmp/kh222_build_stage.db")
    if stage.exists():
        stage.unlink()
    con = sqlite3.connect(stage)
    con.execute("PRAGMA foreign_keys = ON")

    # ----- raw_master_list from the ALL sheet of KH222 Master List.xlsx
    h_ml, r_ml = read_master_list(up / "KH222 Master List.xlsx")
    build_raw_table(con, "raw_master_list", h_ml, r_ml)

    # ----- raw_blue_book / raw_payments / raw_check_ins / raw_gps_48_hours
    csv_sources = {
        "raw_blue_book":     up / "New Blue Book.csv",
        "raw_payments":      up / "Payments (2).csv",
        "raw_check_ins":     up / "Check Ins (1).csv",
        "raw_gps_48_hours":  up / "GPS 48 Hours (1).csv",
    }
    headers = {}
    for tbl, path in csv_sources.items():
        h, r = read_sharepoint_csv(path)
        headers[tbl] = h
        build_raw_table(con, tbl, h, r)

    # Normalized schema
    for stmt in DDL:
        con.execute(stmt)

    # ---------- defendants ----------
    # blue_book first; master_list fills in missing IDNs and last names.
    bb_idx = {h: i for i, h in enumerate(headers["raw_blue_book"])}
    def bb(row, key):
        return row[bb_idx[key]] if key in bb_idx else None

    bb_rows = [list(r) for r in con.execute(
        f'SELECT {",".join(chr(34)+h+chr(34) for h in headers["raw_blue_book"])} FROM raw_blue_book'
    )]

    defendants: dict[int, dict] = {}
    for row in bb_rows:
        idn = parse_int(bb(row, "idn"))
        if idn is None:
            continue
        defendants[idn] = dict(
            idn=idn,
            defendant_name=bb(row, "defendant"),
            defendant_last_name=None,
            birthdate=parse_date(bb(row, "birthdate")),
            pretrial_level=parse_int(bb(row, "pretrial_level")),
            charge_type=bb(row, "charge_type"),
            supervision_type=bb(row, "supervision_type"),
            order_from=bb(row, "order_from"),
            dma=parse_bool(bb(row, "dma")),
            gps=parse_bool(bb(row, "gps")),
            referral_date=parse_date(bb(row, "referral_date")),
            closed_date=parse_date(bb(row, "closed_date")),
            case_status=bb(row, "case_status"),
            supervising_officer=bb(row, "supervising_officer"),
            ptr_successfully_completed=parse_bool(bb(row, "ptr_successfully_completed")),
            bond_amount=parse_money(bb(row, "bond_amount")),
            total_paid=parse_money(bb(row, "total_paid")),
            source="blue_book",
        )

    # Master list (ALL sheet) - 8 columns.
    ml_cols = headers["raw_master_list"]
    ml_idx = {h: i for i, h in enumerate(ml_cols)}
    for row in con.execute(f'SELECT {",".join(chr(34)+c+chr(34) for c in ml_cols)} FROM raw_master_list'):
        idn = parse_int(row[ml_idx.get("idn", -1)] if "idn" in ml_idx else None)
        if idn is None:
            continue
        last = row[ml_idx["defendant_last_name"]] if "defendant_last_name" in ml_idx else None
        pretrial_level = parse_int(row[ml_idx["pretrial_level"]]) if "pretrial_level" in ml_idx else None
        referral_date = parse_date(row[ml_idx["referral_date"]]) if "referral_date" in ml_idx else None
        charge_type = row[ml_idx["charge_type"]] if "charge_type" in ml_idx else None
        order_from = row[ml_idx["order_from"]] if "order_from" in ml_idx else None
        supervision_type = row[ml_idx["supervision_type"]] if "supervision_type" in ml_idx else None
        dma = parse_bool(row[ml_idx["dma"]]) if "dma" in ml_idx else None
        if idn in defendants:
            d = defendants[idn]
            d["source"] = "both"
            if not d.get("defendant_last_name"):
                d["defendant_last_name"] = last
        else:
            defendants[idn] = dict(
                idn=idn, defendant_name=None, defendant_last_name=last,
                birthdate=None, pretrial_level=pretrial_level,
                charge_type=charge_type, supervision_type=supervision_type,
                order_from=order_from, dma=dma, gps=None,
                referral_date=referral_date, closed_date=None,
                case_status=None, supervising_officer=None,
                ptr_successfully_completed=None, bond_amount=None, total_paid=None,
                source="master_list",
            )

    cols = ["idn","defendant_name","defendant_last_name","birthdate","pretrial_level",
            "charge_type","supervision_type","order_from","dma","gps","referral_date",
            "closed_date","case_status","supervising_officer","ptr_successfully_completed",
            "bond_amount","total_paid","source"]
    con.executemany(
        f"INSERT INTO defendants ({','.join(cols)}) VALUES ({','.join('?'*len(cols))})",
        [tuple(d.get(c) for c in cols) for d in defendants.values()],
    )

    # ---------- cases (split @case_num strings) ----------
    case_pairs: set[tuple[int, str]] = set()
    for row in bb_rows:
        idn = parse_int(bb(row, "idn"))
        raw = bb(row, "warrant_case_num")
        if idn is None:
            continue
        for cn in split_cases(raw):
            case_pairs.add((idn, cn))
    # Also split payments.case_number (they sometimes show multi-values).
    pay_cols = headers["raw_payments"]
    pay_idx = {h: i for i, h in enumerate(pay_cols)}
    pay_rows = list(con.execute(f'SELECT {",".join(chr(34)+c+chr(34) for c in pay_cols)} FROM raw_payments'))
    for row in pay_rows:
        idn = parse_int(row[pay_idx["idn"]]) if "idn" in pay_idx else None
        raw = row[pay_idx["case_number"]] if "case_number" in pay_idx else None
        if idn is None:
            continue
        for cn in split_cases(raw):
            case_pairs.add((idn, cn))
    con.executemany(
        "INSERT INTO cases (idn, case_number, source) VALUES (?, ?, ?)",
        [(i, c, "derived") for (i, c) in sorted(case_pairs)],
    )

    # ---------- payments ----------
    con.executemany(
        "INSERT INTO payments (idn, case_number, defendant, payment_date, "
        "payment_amount, officer, payment_type, case_status) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        [(
            parse_int(r[pay_idx["idn"]])                 if "idn" in pay_idx else None,
            r[pay_idx["case_number"]]                     if "case_number" in pay_idx else None,
            r[pay_idx["defendant"]]                       if "defendant" in pay_idx else None,
            parse_date(r[pay_idx["payment_date"]])        if "payment_date" in pay_idx else None,
            parse_money(r[pay_idx["payment_amount"]])     if "payment_amount" in pay_idx else None,
            r[pay_idx.get("officer_that_collected_payment", pay_idx.get("officer", -1))] if ("officer_that_collected_payment" in pay_idx or "officer" in pay_idx) else None,
            r[pay_idx["payment_type"]]                    if "payment_type" in pay_idx else None,
            r[pay_idx["case_status"]]                     if "case_status" in pay_idx else None,
        ) for r in pay_rows],
    )

    # ---------- check_ins ----------
    ci_cols = headers["raw_check_ins"]
    ci_idx = {h: i for i, h in enumerate(ci_cols)}
    ci_rows = list(con.execute(f'SELECT {",".join(chr(34)+c+chr(34) for c in ci_cols)} FROM raw_check_ins'))
    con.executemany(
        "INSERT INTO check_ins (idn, case_number, defendant, check_in_date, "
        "type_of_check_in, supervising_officer, case_status, referral_date, pretrial_level) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [(
            parse_int(r[ci_idx["idn"]])                       if "idn" in ci_idx else None,
            r[ci_idx["case_number"]]                           if "case_number" in ci_idx else None,
            r[ci_idx["defendant"]]                             if "defendant" in ci_idx else None,
            parse_date(r[ci_idx.get("date", ci_idx.get("check_in_date", -1))]) if ("date" in ci_idx or "check_in_date" in ci_idx) else None,
            r[ci_idx["type_of_check_in"]]                      if "type_of_check_in" in ci_idx else None,
            r[ci_idx["supervising_officer"]]                   if "supervising_officer" in ci_idx else None,
            r[ci_idx["case_status"]]                           if "case_status" in ci_idx else None,
            parse_date(r[ci_idx["referral_date"]])             if "referral_date" in ci_idx else None,
            parse_int(r[ci_idx["pretrial_level"]])             if "pretrial_level" in ci_idx else None,
        ) for r in ci_rows],
    )

    # ---------- gps_events ----------
    gp_cols = headers["raw_gps_48_hours"]
    gp_idx = {h: i for i, h in enumerate(gp_cols)}
    gp_rows = list(con.execute(f'SELECT {",".join(chr(34)+c+chr(34) for c in gp_cols)} FROM raw_gps_48_hours'))
    con.executemany(
        "INSERT INTO gps_events (idn, case_number, defendant, referral_date, gps_type, "
        "case_status, paid, victim, victim_idn, victim_time_48, victim_accept_deny_gps, "
        "gps_install_date, court_order, da_emailed, closed_date) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [(
            parse_int(r[gp_idx["idn"]])                         if "idn" in gp_idx else None,
            r[gp_idx["case_number"]]                             if "case_number" in gp_idx else None,
            r[gp_idx["defendant"]]                               if "defendant" in gp_idx else None,
            parse_date(r[gp_idx["referral_date"]])               if "referral_date" in gp_idx else None,
            r[gp_idx["gps_type"]]                                if "gps_type" in gp_idx else None,
            r[gp_idx["case_status"]]                             if "case_status" in gp_idx else None,
            parse_bool(r[gp_idx["paid"]])                        if "paid" in gp_idx else None,
            r[gp_idx["victim"]]                                  if "victim" in gp_idx else None,
            parse_int(r[gp_idx["victim_idn"]])                   if "victim_idn" in gp_idx else None,
            parse_date(r[gp_idx["victim_time_48"]])              if "victim_time_48" in gp_idx else None,
            parse_bool(r[gp_idx["victim_accept_deny_gps"]])      if "victim_accept_deny_gps" in gp_idx else None,
            parse_date(r[gp_idx["gps_install_date"]])            if "gps_install_date" in gp_idx else None,
            r[gp_idx.get("order", gp_idx.get("court_order", -1))] if ("order" in gp_idx or "court_order" in gp_idx) else None,
            parse_bool(r[gp_idx["da_emailed"]])                  if "da_emailed" in gp_idx else None,
            parse_date(r[gp_idx["closed_date"]])                 if "closed_date" in gp_idx else None,
        ) for r in gp_rows],
    )

    con.commit()
    con.close()

    # Byte-copy the finished DB into the OneDrive-mounted outputs folder.
    final = out_dir / "kh222.db"
    # Use truncate+cat rather than os.remove: OneDrive refuses unlink.
    with open(final, "wb") as dst, open(stage, "rb") as src:
        shutil.copyfileobj(src, dst)
    print(f"Wrote {final}")


if __name__ == "__main__":
    main()
